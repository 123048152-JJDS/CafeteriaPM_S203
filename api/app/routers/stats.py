from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, timedelta, date
from typing import List, Optional
import io
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

from app.core.database import get_db
from app.core.security import get_current_user, require_roles
from app.models.sale import Sale
from app.models.order import Order, OrderStatus, OrderDetail
from app.models.product import Product
from app.models.ingredient import Ingredient
from app.models.expense import Expense, Purchase
from app.models.user import User
from app.schemas.stats import DashboardStats, ProductoVendido, VentaPorPeriodo

router = APIRouter()

# ============================================================
# DASHBOARD
# ============================================================
@router.get("/dashboard", response_model=DashboardStats)
def get_dashboard(
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin", "caja"))
):
    hoy = date.today()
    inicio_hoy = datetime(hoy.year, hoy.month, hoy.day)

    ventas_hoy = db.query(func.sum(Sale.monto_total)).filter(
        Sale.fecha >= inicio_hoy
    ).scalar() or 0

    estados_activos = db.query(OrderStatus).filter(
        OrderStatus.nombre.in_(["pendiente", "en_preparacion", "listo", "entregado"])
    ).all()
    ids_activos = [e.id for e in estados_activos]
    pedidos_activos = db.query(Order).filter(Order.id_estado_actual.in_(ids_activos)).count()

    stock_bajo = db.query(Ingredient).filter(
        Ingredient.stock_actual <= Ingredient.stock_minimo
    ).count()

    gastos_hoy = db.query(func.sum(Expense.monto)).filter(
        Expense.fecha == hoy
    ).scalar() or 0

    hace_30_dias = datetime.now() - timedelta(days=30)
    estado_pagado = db.query(OrderStatus).filter(OrderStatus.nombre == "pagado").first()
    top_productos = []
    if estado_pagado:
        top = db.query(
            Product.id,
            Product.nombre,
            func.sum(OrderDetail.cantidad).label("total_vendido")
        ).join(OrderDetail.producto).join(Order).filter(
            Order.id_estado_actual == estado_pagado.id,
            Order.created_at >= hace_30_dias
        ).group_by(Product.id).order_by(func.sum(OrderDetail.cantidad).desc()).limit(5).all()
        top_productos = [
            ProductoVendido(id=p.id, nombre=p.nombre, total_vendido=int(p.total_vendido or 0))
            for p in top
        ]

    return DashboardStats(
        ventas_hoy=float(ventas_hoy),
        pedidos_activos=pedidos_activos,
        stock_bajo=stock_bajo,
        gastos_hoy=float(gastos_hoy),
        productos_mas_vendidos=top_productos
    )

# ============================================================
# RESUMEN DEL MES
# ============================================================
@router.get("/resumen-mes")
def get_resumen_mes(
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin", "caja"))
):
    hoy = date.today()
    primer_dia_mes = hoy.replace(day=1)

    estado_pagado = db.query(OrderStatus).filter(OrderStatus.nombre == "pagado").first()
    if not estado_pagado:
        return {
            "total_ventas": 0.0,
            "total_pedidos": 0,
            "ticket_promedio": 0.0,
            "gastos_totales": 0.0,
            "ganancia_neta": 0.0,
            "fecha_inicio": primer_dia_mes.strftime("%Y-%m-%d"),
            "fecha_fin": hoy.strftime("%Y-%m-%d")
        }

    ventas_mes = db.query(Sale).filter(
        Sale.fecha >= primer_dia_mes,
        Sale.fecha <= hoy
    ).all()

    total_ventas = sum(float(v.monto_total) for v in ventas_mes)
    total_pedidos = len(ventas_mes)
    ticket_promedio = total_ventas / total_pedidos if total_pedidos > 0 else 0.0

    gastos_mes = db.query(Expense).filter(
        Expense.fecha >= primer_dia_mes,
        Expense.fecha <= hoy
    ).all()
    gastos_totales = sum(float(g.monto) for g in gastos_mes)

    ganancia_neta = total_ventas - gastos_totales

    return {
        "total_ventas": round(total_ventas, 2),
        "total_pedidos": total_pedidos,
        "ticket_promedio": round(ticket_promedio, 2),
        "gastos_totales": round(gastos_totales, 2),
        "ganancia_neta": round(ganancia_neta, 2),
        "fecha_inicio": primer_dia_mes.strftime("%Y-%m-%d"),
        "fecha_fin": hoy.strftime("%Y-%m-%d")
    }

# ============================================================
# VENTAS DIARIAS (GRÁFICA)
# ============================================================
@router.get("/ventas-diarias")
def get_ventas_diarias(
    dias: int = 7,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin", "caja"))
):
    hoy = date.today()
    fecha_inicio = hoy - timedelta(days=dias - 1)

    ventas = db.query(Sale).filter(
        Sale.fecha >= fecha_inicio,
        Sale.fecha <= hoy
    ).all()

    ventas_por_dia = {}
    for v in ventas:
        fecha_str = v.fecha.strftime("%Y-%m-%d")
        ventas_por_dia[fecha_str] = ventas_por_dia.get(fecha_str, 0.0) + float(v.monto_total)

    resultado = []
    for i in range(dias):
        fecha = fecha_inicio + timedelta(days=i)
        fecha_str = fecha.strftime("%Y-%m-%d")
        resultado.append({
            "fecha": fecha.strftime("%d/%m"),
            "total": round(ventas_por_dia.get(fecha_str, 0.0), 2)
        })

    return resultado

# ============================================================
# PRODUCTOS MÁS VENDIDOS
# ============================================================
@router.get("/productos-mas-vendidos", response_model=List[ProductoVendido])
def get_productos_mas_vendidos(
    limit: int = 10,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    estado_pagado = db.query(OrderStatus).filter(OrderStatus.nombre == "pagado").first()
    if not estado_pagado:
        return []
    hace_30_dias = datetime.now() - timedelta(days=30)
    results = db.query(
        Product.id,
        Product.nombre,
        func.sum(OrderDetail.cantidad).label("total_vendido")
    ).join(OrderDetail.producto).join(Order).filter(
        Order.id_estado_actual == estado_pagado.id,
        Order.created_at >= hace_30_dias
    ).group_by(Product.id).order_by(func.sum(OrderDetail.cantidad).desc()).limit(limit).all()

    return [
        ProductoVendido(
            id=p.id,
            nombre=p.nombre,
            total_vendido=int(p.total_vendido or 0)
        )
        for p in results
    ]

# ============================================================
# LISTA DE VENTAS (CON FILTROS DE FECHA)
# ============================================================
@router.get("/ventas-lista")
def get_ventas_lista(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin", "caja"))
):
    query = db.query(Sale)
    if fecha_inicio:
        query = query.filter(Sale.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Sale.fecha <= fecha_fin)
    ventas = query.order_by(Sale.fecha.desc()).all()
    return [
        {
            "id": v.id,
            "pedido_id": v.id_pedido,
            "cajero": v.cajero.nombre if v.cajero else "N/A",
            "metodo_pago": v.metodo_pago.nombre if v.metodo_pago else "N/A",
            "monto_total": float(v.monto_total),
            "fecha": v.fecha.isoformat()
        }
        for v in ventas
    ]

# ============================================================
# LISTA DE GASTOS (CON FILTROS DE FECHA)
# ============================================================
@router.get("/gastos-lista")
def get_gastos_lista(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin", "caja"))
):
    query = db.query(Expense)
    if fecha_inicio:
        query = query.filter(Expense.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Expense.fecha <= fecha_fin)
    gastos = query.order_by(Expense.fecha.desc()).all()
    return [
        {
            "id": g.id,
            "descripcion": g.descripcion,
            "monto": float(g.monto),
            "categoria": g.categoria.nombre if g.categoria else "Sin categoría",
            "usuario": g.usuario.nombre if g.usuario else "N/A",
            "fecha": g.fecha.isoformat() if g.fecha else None
        }
        for g in gastos
    ]

# ============================================================
# ESTADÍSTICAS DE PRODUCTOS
# ============================================================
@router.get("/productos-estadisticas")
def get_productos_estadisticas(
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    estado_pagado = db.query(OrderStatus).filter(OrderStatus.nombre == "pagado").first()
    if not estado_pagado:
        return {"mas_vendidos": [], "menos_vendidos": [], "rentabilidad": []}

    mas_vendidos = db.query(
        Product.id,
        Product.nombre,
        Product.precio,
        func.sum(OrderDetail.cantidad).label("total_vendido")
    ).join(OrderDetail.producto).join(Order).filter(
        Order.id_estado_actual == estado_pagado.id
    ).group_by(Product.id).order_by(func.sum(OrderDetail.cantidad).desc()).limit(10).all()

    menos_vendidos = db.query(
        Product.id,
        Product.nombre,
        Product.precio,
        func.sum(OrderDetail.cantidad).label("total_vendido")
    ).join(OrderDetail.producto).join(Order).filter(
        Order.id_estado_actual == estado_pagado.id
    ).group_by(Product.id).order_by(func.sum(OrderDetail.cantidad).asc()).limit(5).all()

    rentabilidad = []
    productos = db.query(Product).all()
    for p in productos:
        costo_ingredientes = 0.0
        for pi in p.ingredientes:
            costo_ingredientes += float(pi.cantidad) * float(pi.ingrediente.costo_unitario)
        margen = float(p.precio) - costo_ingredientes
        rentabilidad.append({
            "id": p.id,
            "nombre": p.nombre,
            "precio": float(p.precio),
            "costo": round(costo_ingredientes, 2),
            "margen": round(margen, 2)
        })
    rentabilidad = sorted(rentabilidad, key=lambda x: x["margen"], reverse=True)[:10]

    return {
        "mas_vendidos": [
            {"id": p.id, "nombre": p.nombre, "total_vendido": int(p.total_vendido or 0)}
            for p in mas_vendidos
        ],
        "menos_vendidos": [
            {"id": p.id, "nombre": p.nombre, "total_vendido": int(p.total_vendido or 0)}
            for p in menos_vendidos
        ],
        "rentabilidad": rentabilidad
    }

# ============================================================
# ESTADO DEL INVENTARIO
# ============================================================
@router.get("/inventario-estado")
def get_inventario_estado(
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin", "cocina"))
):
    ingredientes = db.query(Ingredient).all()
    total_items = len(ingredientes)
    criticos = db.query(Ingredient).filter(
        Ingredient.stock_actual <= Ingredient.stock_minimo
    ).count()

    ultimas_compras = db.query(Purchase).order_by(
        Purchase.created_at.desc()
    ).limit(10).all()

    movimientos = [
        {
            "ingrediente": c.ingrediente.nombre if c.ingrediente else "N/A",
            "cantidad": float(c.cantidad),
            "costo_total": float(c.costo_total),
            "fecha": c.created_at.isoformat() if c.created_at else None
        }
        for c in ultimas_compras
    ]

    return {
        "total_items": total_items,
        "criticos": criticos,
        "ingredientes": [
            {
                "id": i.id,
                "nombre": i.nombre,
                "unidad": i.unidad,
                "stock_actual": float(i.stock_actual),
                "stock_minimo": float(i.stock_minimo),
                "estado": "Crítico" if i.stock_actual <= i.stock_minimo else "Ok"
            }
            for i in ingredientes
        ],
        "movimientos": movimientos
    }

# ============================================================
# HISTORIAL DE ACTIVIDAD
# ============================================================
@router.get("/historial-actividad")
def get_historial_actividad(
    limite: int = 20,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    pedidos_recientes = db.query(Order).order_by(
        Order.created_at.desc()
    ).limit(limite).all()

    ventas_recientes = db.query(Sale).order_by(
        Sale.fecha.desc()
    ).limit(limite).all()

    gastos_recientes = db.query(Expense).order_by(
        Expense.fecha.desc()
    ).limit(limite).all()

    return {
        "pedidos": [
            {
                "id": p.id,
                "mesa": p.mesa.numero if p.mesa else "N/A",
                "estado": p.estado_actual.nombre if p.estado_actual else "N/A",
                "created_at": p.created_at.isoformat() if p.created_at else None
            }
            for p in pedidos_recientes
        ],
        "ventas": [
            {
                "id": v.id,
                "pedido_id": v.id_pedido,
                "monto_total": float(v.monto_total),
                "fecha": v.fecha.isoformat() if v.fecha else None
            }
            for v in ventas_recientes
        ],
        "gastos": [
            {
                "id": g.id,
                "descripcion": g.descripcion,
                "monto": float(g.monto),
                "fecha": g.fecha.isoformat() if g.fecha else None
            }
            for g in gastos_recientes
        ]
    }

# ============================================================
# GASTOS DIARIOS (GRÁFICA)
# ============================================================
@router.get("/gastos-diarios")
def get_gastos_diarios(
    dias: int = 7,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin", "caja"))
):
    hoy = date.today()
    fecha_inicio = hoy - timedelta(days=dias - 1)

    gastos = db.query(Expense).filter(
        Expense.fecha >= fecha_inicio,
        Expense.fecha <= hoy
    ).all()

    gastos_por_dia = {}
    for g in gastos:
        fecha_str = g.fecha.strftime("%Y-%m-%d")
        gastos_por_dia[fecha_str] = gastos_por_dia.get(fecha_str, 0.0) + float(g.monto)

    resultado = []
    for i in range(dias):
        fecha = fecha_inicio + timedelta(days=i)
        fecha_str = fecha.strftime("%Y-%m-%d")
        resultado.append({
            "fecha": fecha.strftime("%d/%m"),
            "total": round(gastos_por_dia.get(fecha_str, 0.0), 2)
        })

    return resultado

# ============================================================
# REPORTE DE VENTAS (PDF) - CON FILTROS DE FECHA
# ============================================================
@router.get("/reporte-ventas/pdf")
def reporte_ventas_pdf(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    flowables = []

    title = Paragraph("Reporte de Ventas - Cafetería", styles['Title'])
    flowables.append(title)
    flowables.append(Spacer(1, 0.2*inch))

    hoy = datetime.now().strftime("%Y-%m-%d %H:%M")
    fecha = Paragraph(f"Generado: {hoy}", styles['Normal'])
    flowables.append(fecha)
    flowables.append(Spacer(1, 0.2*inch))

    # Filtros aplicados
    filtros_texto = f"Fechas: {fecha_inicio or 'sin inicio'} - {fecha_fin or 'sin fin'}"
    filtros_parrafo = Paragraph(f"<i>{filtros_texto}</i>", styles['Normal'])
    flowables.append(filtros_parrafo)
    flowables.append(Spacer(1, 0.1*inch))

    query = db.query(Sale)
    if fecha_inicio:
        query = query.filter(Sale.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Sale.fecha <= fecha_fin)

    ventas = query.order_by(Sale.fecha.desc()).all()

    data = [["ID Venta", "Pedido", "Cajero", "Método", "Monto", "Fecha"]]
    total = 0
    for v in ventas:
        data.append([
            str(v.id),
            str(v.id_pedido),
            v.cajero.nombre if v.cajero else "N/A",
            v.metodo_pago.nombre if v.metodo_pago else "N/A",
            f"${float(v.monto_total):.2f}",
            v.fecha.strftime("%Y-%m-%d %H:%M")
        ])
        total += float(v.monto_total)
    data.append(["", "", "", "TOTAL", f"${total:.2f}", ""])

    tabla = Table(data, colWidths=[0.8*inch]*6)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
        ('GRID', (0,0), (-1,-2), 1, colors.black),
    ]))
    flowables.append(tabla)
    flowables.append(Spacer(1, 0.2*inch))

    total_text = Paragraph(f"<b>Total recaudado (según filtros): ${total:.2f}</b>", styles['Normal'])
    flowables.append(total_text)

    doc.build(flowables)
    buffer.seek(0)
    return Response(buffer.getvalue(), media_type="application/pdf", headers={
        "Content-Disposition": "attachment; filename=reporte_ventas.pdf"
    })

# ============================================================
# REPORTE DE VENTAS (XLSX) - CON FILTROS DE FECHA
# ============================================================
@router.get("/reporte-ventas/xlsx")
def reporte_ventas_xlsx(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    headers = ["ID Venta", "Pedido", "Cajero", "Método Pago", "Monto", "Fecha"]
    ws.append(headers)

    query = db.query(Sale)
    if fecha_inicio:
        query = query.filter(Sale.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Sale.fecha <= fecha_fin)

    total = 0
    for v in query.order_by(Sale.fecha.desc()).all():
        total += float(v.monto_total)
        ws.append([
            v.id,
            v.id_pedido,
            v.cajero.nombre if v.cajero else "N/A",
            v.metodo_pago.nombre if v.metodo_pago else "N/A",
            float(v.monto_total),
            v.fecha.strftime("%Y-%m-%d %H:%M")
        ])
    ws.append([])
    ws.append(["TOTAL", "", "", "", total, ""])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": "attachment; filename=reporte_ventas.xlsx"
    })

# ============================================================
# REPORTE DE GASTOS (PDF) - CON FILTROS DE FECHA
# ============================================================
@router.get("/reporte-gastos/pdf")
def reporte_gastos_pdf(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    flowables = []

    title = Paragraph("Reporte de Gastos - Cafetería", styles['Title'])
    flowables.append(title)
    flowables.append(Spacer(1, 0.2*inch))

    hoy = datetime.now().strftime("%Y-%m-%d %H:%M")
    fecha = Paragraph(f"Generado: {hoy}", styles['Normal'])
    flowables.append(fecha)
    flowables.append(Spacer(1, 0.2*inch))

    # Filtros aplicados
    filtros_texto = f"Fechas: {fecha_inicio or 'sin inicio'} - {fecha_fin or 'sin fin'}"
    filtros_parrafo = Paragraph(f"<i>{filtros_texto}</i>", styles['Normal'])
    flowables.append(filtros_parrafo)
    flowables.append(Spacer(1, 0.1*inch))

    query = db.query(Expense)
    if fecha_inicio:
        query = query.filter(Expense.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Expense.fecha <= fecha_fin)

    gastos = query.order_by(Expense.fecha.desc()).all()

    data = [["ID", "Descripción", "Categoría", "Usuario", "Monto", "Fecha"]]
    total = 0
    for g in gastos:
        total += float(g.monto)
        data.append([
            str(g.id),
            g.descripcion,
            g.categoria.nombre if g.categoria else "N/A",
            g.usuario.nombre if g.usuario else "N/A",
            f"${float(g.monto):.2f}",
            g.fecha.strftime("%Y-%m-%d") if g.fecha else ""
        ])
    data.append(["", "", "", "TOTAL", f"${total:.2f}", ""])

    tabla = Table(data, colWidths=[0.8*inch, 2*inch, 1.5*inch, 1.5*inch, 1.2*inch, 1.2*inch])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-2), 1, colors.black),
        ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
    ]))
    flowables.append(tabla)

    doc.build(flowables)
    buffer.seek(0)
    return Response(buffer.getvalue(), media_type="application/pdf", headers={
        "Content-Disposition": "attachment; filename=reporte_gastos.pdf"
    })

# ============================================================
# REPORTE DE GASTOS (XLSX) - CON FILTROS DE FECHA
# ============================================================
@router.get("/reporte-gastos/xlsx")
def reporte_gastos_xlsx(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos"

    headers = ["ID", "Descripción", "Categoría", "Usuario", "Monto", "Fecha"]
    ws.append(headers)

    query = db.query(Expense)
    if fecha_inicio:
        query = query.filter(Expense.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Expense.fecha <= fecha_fin)

    total = 0
    for g in query.order_by(Expense.fecha.desc()).all():
        total += float(g.monto)
        ws.append([
            g.id,
            g.descripcion,
            g.categoria.nombre if g.categoria else "N/A",
            g.usuario.nombre if g.usuario else "N/A",
            float(g.monto),
            g.fecha.strftime("%Y-%m-%d") if g.fecha else ""
        ])
    ws.append([])
    ws.append(["TOTAL", "", "", "", total, ""])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": "attachment; filename=reporte_gastos.xlsx"
    })

# ============================================================
# REPORTE DE PRODUCTOS (PDF)
# ============================================================
@router.get("/reporte-productos/pdf")
def reporte_productos_pdf(
    categoria_id: Optional[int] = None,
    disponible: Optional[bool] = None,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    flowables = []

    title = Paragraph("Reporte de Productos - Cafetería", styles['Title'])
    flowables.append(title)
    flowables.append(Spacer(1, 0.2*inch))

    hoy = datetime.now().strftime("%Y-%m-%d %H:%M")
    fecha = Paragraph(f"Generado: {hoy}", styles['Normal'])
    flowables.append(fecha)
    flowables.append(Spacer(1, 0.2*inch))

    query = db.query(Product)
    if categoria_id:
        query = query.filter(Product.id_categoria == categoria_id)
    if disponible is not None:
        query = query.filter(Product.disponible == disponible)

    productos = query.all()

    data = [["ID", "Nombre", "Categoría", "Precio", "Disponible"]]
    for p in productos:
        data.append([
            str(p.id),
            p.nombre,
            p.categoria.nombre if p.categoria else "N/A",
            f"${float(p.precio):.2f}",
            "Sí" if p.disponible else "No"
        ])

    tabla = Table(data, colWidths=[0.8*inch, 2*inch, 1.5*inch, 1.2*inch, 1.2*inch])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    flowables.append(tabla)

    doc.build(flowables)
    buffer.seek(0)
    return Response(buffer.getvalue(), media_type="application/pdf", headers={
        "Content-Disposition": "attachment; filename=reporte_productos.pdf"
    })

# ============================================================
# REPORTE DE PRODUCTOS (XLSX)
# ============================================================
@router.get("/reporte-productos/xlsx")
def reporte_productos_xlsx(
    categoria_id: Optional[int] = None,
    disponible: Optional[bool] = None,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = ["ID", "Nombre", "Categoría", "Precio", "Disponible"]
    ws.append(headers)

    query = db.query(Product)
    if categoria_id:
        query = query.filter(Product.id_categoria == categoria_id)
    if disponible is not None:
        query = query.filter(Product.disponible == disponible)

    for p in query.all():
        ws.append([
            p.id,
            p.nombre,
            p.categoria.nombre if p.categoria else "N/A",
            float(p.precio),
            "Sí" if p.disponible else "No"
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": "attachment; filename=reporte_productos.xlsx"
    })

# ============================================================
# REPORTE DE PEDIDOS (PDF)
# ============================================================
@router.get("/reporte-pedidos/pdf")
def reporte_pedidos_pdf(
    estado_id: Optional[int] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    flowables = []

    title = Paragraph("Reporte de Pedidos - Cafetería", styles['Title'])
    flowables.append(title)
    flowables.append(Spacer(1, 0.2*inch))

    hoy = datetime.now().strftime("%Y-%m-%d %H:%M")
    fecha = Paragraph(f"Generado: {hoy}", styles['Normal'])
    flowables.append(fecha)
    flowables.append(Spacer(1, 0.2*inch))

    query = db.query(Order)
    if estado_id:
        query = query.filter(Order.id_estado_actual == estado_id)
    if fecha_inicio:
        query = query.filter(Order.created_at >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Order.created_at <= fecha_fin)

    pedidos = query.order_by(Order.created_at.desc()).all()

    data = [["ID", "Mesa", "Mesero", "Estado", "Total", "Fecha"]]
    for p in pedidos:
        total = sum(float(d.subtotal or 0) for d in p.detalles)
        data.append([
            str(p.id),
            f"Mesa {p.mesa.numero}" if p.mesa else "N/A",
            p.mesero.nombre if p.mesero else "N/A",
            p.estado_actual.nombre if p.estado_actual else "N/A",
            f"${total:.2f}",
            p.created_at.strftime("%Y-%m-%d %H:%M")
        ])

    tabla = Table(data, colWidths=[0.8*inch]*6)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    flowables.append(tabla)

    doc.build(flowables)
    buffer.seek(0)
    return Response(buffer.getvalue(), media_type="application/pdf", headers={
        "Content-Disposition": "attachment; filename=reporte_pedidos.pdf"
    })

# ============================================================
# REPORTE DE PEDIDOS (XLSX)
# ============================================================
@router.get("/reporte-pedidos/xlsx")
def reporte_pedidos_xlsx(
    estado_id: Optional[int] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    headers = ["ID", "Mesa", "Mesero", "Estado", "Total", "Fecha"]
    ws.append(headers)

    query = db.query(Order)
    if estado_id:
        query = query.filter(Order.id_estado_actual == estado_id)
    if fecha_inicio:
        query = query.filter(Order.created_at >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Order.created_at <= fecha_fin)

    for p in query.order_by(Order.created_at.desc()).all():
        total = sum(float(d.subtotal or 0) for d in p.detalles)
        ws.append([
            p.id,
            f"Mesa {p.mesa.numero}" if p.mesa else "N/A",
            p.mesero.nombre if p.mesero else "N/A",
            p.estado_actual.nombre if p.estado_actual else "N/A",
            total,
            p.created_at.strftime("%Y-%m-%d %H:%M")
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": "attachment; filename=reporte_pedidos.xlsx"
    })

# ============================================================
# REPORTE DE INVENTARIO (PDF)
# ============================================================
@router.get("/reporte-inventario/pdf")
def reporte_inventario_pdf(
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    flowables = []

    title = Paragraph("Reporte de Inventario - Cafetería", styles['Title'])
    flowables.append(title)
    flowables.append(Spacer(1, 0.2*inch))

    hoy = datetime.now().strftime("%Y-%m-%d %H:%M")
    fecha = Paragraph(f"Generado: {hoy}", styles['Normal'])
    flowables.append(fecha)
    flowables.append(Spacer(1, 0.2*inch))

    ingredientes = db.query(Ingredient).all()

    data = [["ID", "Nombre", "Unidad", "Stock actual", "Stock mínimo", "Estado"]]
    for i in ingredientes:
        estado = "Crítico" if i.stock_actual <= i.stock_minimo else "Ok"
        data.append([
            str(i.id),
            i.nombre,
            i.unidad,
            f"{float(i.stock_actual):.1f}",
            f"{float(i.stock_minimo):.1f}",
            estado
        ])

    tabla = Table(data, colWidths=[0.8*inch, 2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    flowables.append(tabla)

    doc.build(flowables)
    buffer.seek(0)
    return Response(buffer.getvalue(), media_type="application/pdf", headers={
        "Content-Disposition": "attachment; filename=reporte_inventario.pdf"
    })

# ============================================================
# REPORTE DE INVENTARIO (XLSX)
# ============================================================
@router.get("/reporte-inventario/xlsx")
def reporte_inventario_xlsx(
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = ["ID", "Nombre", "Unidad", "Stock actual", "Stock mínimo", "Estado"]
    ws.append(headers)

    for i in db.query(Ingredient).all():
        estado = "Crítico" if i.stock_actual <= i.stock_minimo else "Ok"
        ws.append([
            i.id,
            i.nombre,
            i.unidad,
            float(i.stock_actual),
            float(i.stock_minimo),
            estado
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": "attachment; filename=reporte_inventario.xlsx"
    })

# ============================================================
# REPORTE DE HISTORIAL (PDF)
# ============================================================
@router.get("/reporte-historial/pdf")
def reporte_historial_pdf(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    flowables = []

    title = Paragraph("Reporte de Historial - Cafetería", styles['Title'])
    flowables.append(title)
    flowables.append(Spacer(1, 0.2*inch))

    hoy = datetime.now().strftime("%Y-%m-%d %H:%M")
    fecha = Paragraph(f"Generado: {hoy}", styles['Normal'])
    flowables.append(fecha)
    flowables.append(Spacer(1, 0.2*inch))

    # Filtros aplicados
    filtros_texto = f"Fechas: {fecha_inicio or 'sin inicio'} - {fecha_fin or 'sin fin'}"
    filtros_parrafo = Paragraph(f"<i>{filtros_texto}</i>", styles['Normal'])
    flowables.append(filtros_parrafo)
    flowables.append(Spacer(1, 0.1*inch))

    # Pedidos recientes
    flowables.append(Paragraph("<b>Pedidos Recientes</b>", styles['Normal']))
    query = db.query(Order)
    if fecha_inicio:
        query = query.filter(Order.created_at >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Order.created_at <= fecha_fin)
    pedidos = query.order_by(Order.created_at.desc()).limit(20).all()

    data_pedidos = [["ID", "Mesa", "Estado", "Fecha"]]
    for p in pedidos:
        data_pedidos.append([
            str(p.id),
            f"Mesa {p.mesa.numero}" if p.mesa else "N/A",
            p.estado_actual.nombre if p.estado_actual else "N/A",
            p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else ""
        ])
    tabla_pedidos = Table(data_pedidos, colWidths=[0.8*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    tabla_pedidos.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    flowables.append(tabla_pedidos)
    flowables.append(Spacer(1, 0.2*inch))

    # Ventas recientes
    flowables.append(Paragraph("<b>Ventas Recientes</b>", styles['Normal']))
    query_ventas = db.query(Sale)
    if fecha_inicio:
        query_ventas = query_ventas.filter(Sale.fecha >= fecha_inicio)
    if fecha_fin:
        query_ventas = query_ventas.filter(Sale.fecha <= fecha_fin)
    ventas = query_ventas.order_by(Sale.fecha.desc()).limit(20).all()

    data_ventas = [["ID", "Pedido", "Monto", "Fecha"]]
    for v in ventas:
        data_ventas.append([
            str(v.id),
            str(v.id_pedido),
            f"${float(v.monto_total):.2f}",
            v.fecha.strftime("%Y-%m-%d %H:%M") if v.fecha else ""
        ])
    tabla_ventas = Table(data_ventas, colWidths=[0.8*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    tabla_ventas.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    flowables.append(tabla_ventas)
    flowables.append(Spacer(1, 0.2*inch))

    # Gastos recientes
    flowables.append(Paragraph("<b>Gastos Recientes</b>", styles['Normal']))
    query_gastos = db.query(Expense)
    if fecha_inicio:
        query_gastos = query_gastos.filter(Expense.fecha >= fecha_inicio)
    if fecha_fin:
        query_gastos = query_gastos.filter(Expense.fecha <= fecha_fin)
    gastos = query_gastos.order_by(Expense.fecha.desc()).limit(20).all()

    data_gastos = [["ID", "Descripción", "Monto", "Fecha"]]
    for g in gastos:
        data_gastos.append([
            str(g.id),
            g.descripcion,
            f"${float(g.monto):.2f}",
            g.fecha.strftime("%Y-%m-%d") if g.fecha else ""
        ])
    tabla_gastos = Table(data_gastos, colWidths=[0.8*inch, 2*inch, 1.5*inch, 1.5*inch])
    tabla_gastos.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    flowables.append(tabla_gastos)

    doc.build(flowables)
    buffer.seek(0)
    return Response(buffer.getvalue(), media_type="application/pdf", headers={
        "Content-Disposition": "attachment; filename=reporte_historial.pdf"
    })

# ============================================================
# REPORTE DE HISTORIAL (XLSX)
# ============================================================
@router.get("/reporte-historial/xlsx")
def reporte_historial_xlsx(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin"))
):
    wb = openpyxl.Workbook()
    ws_pedidos = wb.active
    ws_pedidos.title = "Pedidos"

    # Hoja de Pedidos
    headers = ["ID", "Mesa", "Estado", "Fecha"]
    ws_pedidos.append(headers)
    query = db.query(Order)
    if fecha_inicio:
        query = query.filter(Order.created_at >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Order.created_at <= fecha_fin)
    for p in query.order_by(Order.created_at.desc()).limit(20).all():
        ws_pedidos.append([
            p.id,
            f"Mesa {p.mesa.numero}" if p.mesa else "N/A",
            p.estado_actual.nombre if p.estado_actual else "N/A",
            p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else ""
        ])

    # Hoja de Ventas
    ws_ventas = wb.create_sheet("Ventas")
    headers_ventas = ["ID", "Pedido", "Monto", "Fecha"]
    ws_ventas.append(headers_ventas)
    query_ventas = db.query(Sale)
    if fecha_inicio:
        query_ventas = query_ventas.filter(Sale.fecha >= fecha_inicio)
    if fecha_fin:
        query_ventas = query_ventas.filter(Sale.fecha <= fecha_fin)
    for v in query_ventas.order_by(Sale.fecha.desc()).limit(20).all():
        ws_ventas.append([
            v.id,
            v.id_pedido,
            float(v.monto_total),
            v.fecha.strftime("%Y-%m-%d %H:%M") if v.fecha else ""
        ])

    # Hoja de Gastos
    ws_gastos = wb.create_sheet("Gastos")
    headers_gastos = ["ID", "Descripción", "Monto", "Fecha"]
    ws_gastos.append(headers_gastos)
    query_gastos = db.query(Expense)
    if fecha_inicio:
        query_gastos = query_gastos.filter(Expense.fecha >= fecha_inicio)
    if fecha_fin:
        query_gastos = query_gastos.filter(Expense.fecha <= fecha_fin)
    for g in query_gastos.order_by(Expense.fecha.desc()).limit(20).all():
        ws_gastos.append([
            g.id,
            g.descripcion,
            float(g.monto),
            g.fecha.strftime("%Y-%m-%d") if g.fecha else ""
        ])

    for sheet in [ws_pedidos, ws_ventas, ws_gastos]:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            sheet.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": "attachment; filename=reporte_historial.xlsx"
    })